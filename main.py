"""
Track the stock performance on the stock market and store price development.
Trigger notifications to the user in case of high market volatility.
"""

from bs4 import BeautifulSoup as html
from pprint import pprint
import openpyxl
import os
import pandas
import requests
import urllib
import xlrd


class Stock():

    URL_DATA = "https://www.tradegate.de"
    URL_QUERY = URL_DATA + "/orderbuch.php?isin="
    FILE_DATA = "performance.xlsx"

    def __init__(self, id):
        self._id = id
        self._query = Stock.URL_QUERY + id
        self._html = html(requests.get(self._query).content, "html.parser")
        self._data = {}
        self._image = {}
        self._fetchData()
        self._fetchImage()

    def _fetchData(self):
        stock_data = {}
        stock_name = self._html.find("div", {"id": "col1_content"}).find("h2").text
        stock_data.update({"name": stock_name})
        tag_parent = self._html.find_all("td", class_="longprice")
        for item in tag_parent:
            if "id" in item.attrs:
                key = item.attrs.get("id")
                value = item.text.replace(" ", "").replace(",", ".").replace("\xa0", "").replace("TEUR", " TEUR")
                stock_data.update({key: value})
            else:
                tag_child = item.find("strong")
                key = tag_child.attrs.get("id")
                value = item.text.replace(" ", "").replace(",", ".").replace("\xa0", "").replace("TEUR", " TEUR")
                stock_data.update({key: value})
        self._data = stock_data

    def _fetchImage(self):
        stock_intraday = self._html.find("img", {"id": "intraday"}).attrs
        stock_week = self._html.find("img", {"id": "woche"}).attrs
        stock_month = self._html.find("img", {"id": "monat"}).attrs
        stock_month6 = self._html.find("img", {"id": "monat6"}).attrs
        stock_year = self._html.find("img", {"id": "jahr"}).attrs
        self._image = {"intraday": stock_intraday,
            "week": stock_week,
            "month": stock_month,
            "month6": stock_month6,
            "year": stock_year}

    def getData(self, key = "all"):
        if (key == "all"):
            return self._data
        else:
            return self._data[key]

    def getImage(self, key = "all"):
        if (key == "all"):
            for key, _value in self._image.items():
                self.getImage(key)
        else:
            image_name = self._image[key]["alt"].split(" - ")[-1].replace(" ", "_").lower() + ".png"
            image_url = Stock.URL_DATA + self._image[key]["src"]
            image_path = os.getcwd() + "/misc/" + self.getData("name").split()[0].lower() + "/"
            if not os.path.exists(image_path):
                os.makedirs(image_path)
            urllib.request.urlretrieve(image_url, image_path + image_name)

    def saveData(self):
        data_file = Stock.FILE_DATA
        data_sheet = self.getData("name").split()[0]
        data_path = os.getcwd() + "/misc/"
        data_file = data_path + data_file
        data = [pandas.to_datetime("today").strftime("%Y-%m-%d  %H:%M"),
            float(self._data["last"]),
            float(self._data["low"]),
            float(self._data["high"]),
            self._data["delta"],
            float(self._data["preis"])]
        try:
            writer = pandas.ExcelWriter(data_file, engine="openpyxl")
            df = pandas.DataFrame(data).T
            df.columns = ["Date", "Last", "Low", "High", "Change", "Mean"]
            if os.path.exists(data_file):
                reader = pandas.ExcelFile(data_file)
                if data_sheet in reader.sheet_names:
                    df = reader.parse(data_sheet)
                    df.loc[-1] = data
                reader.close()
                workbook = openpyxl.load_workbook(data_file)
                writer.book = workbook
                writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            else:
                workbook = openpyxl.Workbook()
            df.to_excel(writer, sheet_name=data_sheet, startrow=0, index=False, header=True)
            worksheet = writer.book[data_sheet]
            worksheet.column_dimensions["A"].width = 17
            writer.close()
        except ValueError:
            print("Error@saveData: Could not write data!")
        except xlrd.biffh.XLRDError:
            print("Error@saveData: Invalid sheet data format!")
        finally:
            pass

    def saveImage(self):
        data_file = Stock.FILE_DATA
        data_sheet = "Performance"
        data_path = os.getcwd() + "/misc/"
        data_file = data_path + data_file
        data_name = self.getData("name").split()[0]
        alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        try:
            writer = pandas.ExcelWriter(data_file, engine="openpyxl")
            if os.path.exists(data_file):
                reader = pandas.ExcelFile(data_file)
                if data_sheet in reader.sheet_names:
                    df = reader.parse(data_sheet, header=None)
                    index_row = df.loc[0, 52]
                else:
                    index_row = 0
                reader.close()
                workbook = openpyxl.load_workbook(data_file)
                writer.book = workbook
                writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            else:
                workbook = openpyxl.Workbook()
            df = pandas.DataFrame({"name": [data_name]})
            df.to_excel(writer, sheet_name=data_sheet, startrow=index_row, startcol=0, index=False, header=False)
            df = pandas.DataFrame({"index": [index_row + 13]})
            df.to_excel(writer, sheet_name=data_sheet, startrow=0, startcol=52, index=False, header=False)
            index_column = 0
            for image_file in os.listdir(data_path + data_name):
                if image_file.endswith(".png"):
                    img = openpyxl.drawing.image.Image(os.path.join(data_path + data_name, image_file))
                    img.anchor = alphabet[index_column] + str(index_row + 2)
                    index_column = index_column + 6
                    workbook[data_sheet].add_image(img)
                    workbook.save(data_file)
                else:
                    continue
            # workbook.remove(workbook[data_sheet])
            # workbook.save(data_file)
            writer.close()
        except ValueError:
            print("Error@saveImage: Could not write data!")
        except xlrd.biffh.XLRDError:
            print("Error@saveImage: Invalid sheet data format!")
        finally:
            pass


def main():
    # Open connection to stock exchange database
    amazon = Stock("US0231351067")
    microsoft = Stock("US5949181045")
    apple = Stock("US0378331005")
    google = Stock("US02079K3059")
    # Retrieve stock performance data
    pprint(amazon.getData())
    pprint(microsoft.getData())
    pprint(apple.getData())
    pprint(google.getData())
    # Download stock performance images
    amazon.getImage()
    microsoft.getImage()
    apple.getImage()
    google.getImage()
    # Save stock performance data
    amazon.saveData()
    microsoft.saveData()
    apple.saveData()
    google.saveData()
    # Save stock performance images
    amazon.saveImage()
    microsoft.saveImage()
    # apple.saveImage()
    # google.saveImage()

if __name__ == "__main__":
    main()
