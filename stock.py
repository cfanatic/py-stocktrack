from bs4 import BeautifulSoup as html
import openpyxl
import os
import pandas
import requests
import shutil
import urllib
import xlrd


class Stock():

    URL_DATA = "https://www.tradegate.de"
    URL_QUERY = URL_DATA + "/orderbuch.php?isin="
    FILE_DATA = "performance.xlsx"
    FILE_TAB = "Performance"

    def __init__(self, id):
        self._id = id
        self._query = Stock.URL_QUERY + id
        self._html = html(requests.get(self._query).content, "html.parser")
        self._data = {}
        self._image = {}
        self._fetchData()
        self._fetchImage()

    def _fetchData(self):
        stock_data = {}
        stock_name = self._html.find(
            "div", {"id": "col1_content"}).find("h2").text
        stock_data.update({"name": stock_name})
        tag_parent = self._html.find_all("td", class_="longprice")
        for item in tag_parent:
            if "id" in item.attrs:
                key = item.attrs.get("id")
                value = item.text.replace(" ", "").replace(
                    ",", ".").replace("\xa0", "").replace("TEUR", " TEUR")
                stock_data.update({key: value})
            else:
                tag_child = item.find("strong")
                key = tag_child.attrs.get("id")
                value = item.text.replace(" ", "").replace(
                    ",", ".").replace("\xa0", "").replace("TEUR", " TEUR")
                stock_data.update({key: value})
        self._data = stock_data

    def _fetchImage(self):
        stock_intraday = self._html.find("img", {"id": "intraday"}).attrs
        stock_week = self._html.find("img", {"id": "woche"}).attrs
        stock_month = self._html.find("img", {"id": "monat"}).attrs
        stock_month6 = self._html.find("img", {"id": "monat6"}).attrs
        stock_year = self._html.find("img", {"id": "jahr"}).attrs
        self._image = {"intraday": stock_intraday,
                       "week": stock_week,
                       "month": stock_month,
                       "month6": stock_month6,
                       "year": stock_year}

    def getData(self, key="all"):
        if (key == "all"):
            return self._data
        else:
            return self._data[key]

    def getImage(self, key="all"):
        if (key == "all"):
            for key, _value in self._image.items():
                self.getImage(key)
        else:
            image_name = self._image[key]["alt"].split(
                " - ")[-1].replace(" ", "_").lower() + ".png"
            image_url = Stock.URL_DATA + self._image[key]["src"]
            image_path = os.path.join(
                os.getcwd(), "misc", self.getData("name").split()[0].lower())
            if not os.path.exists(image_path):
                os.makedirs(image_path)
            urllib.request.urlretrieve(
                image_url, os.path.join(image_path, image_name))

    def saveData(self):
        data_file = Stock.FILE_DATA
        data_sheet = self.getData("name").split()[0]
        data_path = os.path.join(os.getcwd(), "misc")
        data_file = os.path.join(data_path, data_file)
        data_backup = data_file + ".backup"
        data = [pandas.to_datetime("today").strftime("%Y-%m-%d  %H:%M"),
                float(self._data["last"]),
                float(self._data["low"]),
                float(self._data["high"]),
                self._data["delta"],
                float(self._data["preis"])]
        try:
            writer = pandas.ExcelWriter(data_file, engine="openpyxl")
            df = pandas.DataFrame(data).T
            df.columns = ["Date", "Last", "Low", "High", "Change", "Mean"]
            if os.path.exists(data_file):
                shutil.copyfile(data_file, data_backup)
                reader = pandas.ExcelFile(data_file)
                if data_sheet in reader.sheet_names:
                    df = reader.parse(data_sheet)
                    df.loc[-1] = data
                reader.close()
                workbook = openpyxl.load_workbook(data_file)
                writer.book = workbook
                writer.sheets = dict((ws.title, ws)
                                     for ws in workbook.worksheets)
            else:
                workbook = openpyxl.Workbook()
            df.to_excel(writer, sheet_name=data_sheet,
                        startrow=0, index=False, header=True)
            worksheet = writer.book[data_sheet]
            worksheet.column_dimensions["A"].width = 17
            cell_last = "B" + str(worksheet.max_row)
            cell_low = "C" + str(worksheet.max_row)
            cell_high = "D" + str(worksheet.max_row)
            cell_change = "E" + str(worksheet.max_row)
            cell_mean = "F" + str(worksheet.max_row)
            worksheet[cell_last].number_format = "#,##0.00€"
            worksheet[cell_low].number_format = "#,##0.00€"
            worksheet[cell_high].number_format = "#,##0.00€"
            worksheet[cell_mean].number_format = "#,##0.00€"
            worksheet[cell_change].value = float(
                worksheet[cell_change].value.replace("%", "")) / 100.0
            worksheet[cell_change].number_format = "0.00%"
            writer.close()
        except FileNotFoundError as e:
            print("Error@saveData: File does not exist: " + e.filename)
        except ValueError:
            print("Error@saveImage: " + e.args[0])
        except xlrd.biffh.XLRDError:
            print("Error@saveData: Invalid sheet data format!")
        else:
            if os.path.exists(data_backup):
                os.remove(data_backup)
        finally:
            if os.path.exists(data_backup):
                shutil.move(data_backup, data_file)

    @staticmethod
    def saveImages():
        data_file = Stock.FILE_DATA
        data_sheet = Stock.FILE_TAB
        data_path = os.path.join(os.getcwd(), "misc")
        data_file = os.path.join(data_path, data_file)
        data_backup = data_file + ".backup"
        data_names = next(os.walk(data_path))[1]
        data_row = 0
        try:
            shutil.copyfile(data_file, data_backup)
            writer = pandas.ExcelWriter(data_file, engine="openpyxl")
            workbook = openpyxl.load_workbook(data_file)
            writer.book = workbook
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            for data_name in data_names:
                data_column = 0
                df = pandas.DataFrame({"name": [data_name.capitalize()]})
                df.to_excel(writer, sheet_name=data_sheet,
                            startrow=data_row, startcol=0, index=False, header=False)
                for image_file in ["intraday.png", "1_woche.png", "1_monat.png", "6_monate.png", "1_jahr.png"]:
                    img = openpyxl.drawing.image.Image(
                        os.path.join(data_path, data_name, image_file))
                    img.anchor = openpyxl.utils.get_column_letter(
                        data_column + 1) + str(data_row + 2)
                    data_column += 6
                    workbook[data_sheet].add_image(img)
                    workbook.save(data_file)
                for caption in [(3, "Intraday"), (9, "Week"), (15, "Month"), (21, "Month-6"), (27, "Year")]:
                    cell = openpyxl.utils.get_column_letter(
                        caption[0]) + str(data_row + 1)
                    workbook[data_sheet][cell] = caption[1]
                caption = workbook[data_sheet]["A" + str(data_row + 1)]
                caption.font = openpyxl.styles.Font(bold=True)
                data_row += 13
            writer.close()
        except FileNotFoundError as e:
            print("Error@saveImage: File does not exist: " + e.filename)
        except ValueError as e:
            print("Error@saveImage: " + e.args[0])
        except xlrd.biffh.XLRDError:
            print("Error@saveImage: Invalid sheet data format!")
        else:
            if os.path.exists(data_backup):
                os.remove(data_backup)
        finally:
            if os.path.exists(data_backup):
                shutil.move(data_backup, data_file)

    @staticmethod
    def deleteImages():
        data_file = Stock.FILE_DATA
        data_sheet = Stock.FILE_TAB
        data_path = os.path.join(os.getcwd(), "misc")
        data_file = os.path.join(data_path, data_file)
        data_backup = data_file + ".backup"
        try:
            shutil.copyfile(data_file, data_backup)
            workbook = openpyxl.load_workbook(data_file)
            workbook.remove(workbook[data_sheet])
            workbook.save(data_file)
        except FileNotFoundError as e:
            print("Error@deleteImages: File does not exist: " + e.filename)
        except KeyError as e:
            print("Error@deleteImages: " + e.args[0])
        else:
            if os.path.exists(data_backup):
                os.remove(data_backup)
        finally:
            if os.path.exists(data_backup):
                shutil.move(data_backup, data_file)
